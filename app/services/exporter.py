from __future__ import annotations

import csv
import io
import json
import re
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import asdict, dataclass, is_dataclass
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class ExportColumn:
    key: str
    label: str
    aliases: tuple[str, ...] = ()


DEFAULT_RECRUITMENT_COLUMNS: tuple[ExportColumn, ...] = (
    ExportColumn(
        "candidate_name",
        "候选人",
        ("candidate.name", "candidate_name", "name"),
    ),
    ExportColumn("job_title", "岗位", ("job.title", "job_title", "position")),
    ExportColumn("status", "状态", ("application.status", "status")),
    ExportColumn(
        "total_score",
        "评分",
        ("evaluation.total_score", "total_score", "score"),
    ),
    ExportColumn(
        "owner",
        "负责人",
        ("owner_name", "owner.name", "owner", "owner_id"),
    ),
    ExportColumn(
        "interview_time",
        "面试时间",
        (
            "interview.confirmed_slot",
            "confirmed_slot",
            "interview_time",
            "start_at",
        ),
    ),
    ExportColumn(
        "updated_at",
        "更新时间",
        ("application.updated_at", "updated_at"),
    ),
)


def export_csv(
    records_or_session: Any,
    *,
    columns: Any = None,
    statement: Any = None,
    model: Any = None,
    scalars: bool = True,
    encoding: str = "utf-8-sig",
    line_terminator: str = "\r\n",
) -> bytes:
    """将普通记录或 SQLAlchemy 查询结果导出为 CSV 字节。"""

    records = load_export_records(
        records_or_session,
        statement=statement,
        model=model,
        scalars=scalars,
    )
    specs = _resolve_columns(records, columns)
    output = io.StringIO(newline="")
    writer = csv.writer(output, lineterminator=line_terminator)
    writer.writerow([column.label for column in specs])
    for record in records:
        writer.writerow([_csv_value(_column_value(record, column)) for column in specs])
    return output.getvalue().encode(encoding)


def export_xlsx(
    records_or_session: Any,
    *,
    columns: Any = None,
    statement: Any = None,
    model: Any = None,
    scalars: bool = True,
    sheet_name: str = "招聘数据",
) -> bytes:
    """将普通记录或 SQLAlchemy 查询结果导出为格式化 XLSX 字节。"""

    records = load_export_records(
        records_or_session,
        statement=statement,
        model=model,
        scalars=scalars,
    )
    specs = _resolve_columns(records, columns)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = _safe_sheet_name(sheet_name)
    worksheet.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for column_index, column in enumerate(specs, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=column.label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    max_widths = [len(column.label) for column in specs]
    for row_index, record in enumerate(records, start=2):
        for column_index, column in enumerate(specs, start=1):
            value = _xlsx_value(_column_value(record, column))
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            max_widths[column_index - 1] = max(
                max_widths[column_index - 1],
                _display_width(value),
            )

    if specs:
        worksheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(specs))}{max(1, len(records) + 1)}"
        )
    for index, width in enumerate(max_widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = min(
            max(width + 2, 10), 50
        )
    worksheet.row_dimensions[1].height = 22

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def export_recruitment_csv(records_or_session: Any, **kwargs: Any) -> bytes:
    """按腾讯文档兼容的招聘字段顺序导出 CSV。"""

    kwargs.setdefault("columns", DEFAULT_RECRUITMENT_COLUMNS)
    return export_csv(records_or_session, **kwargs)


def export_recruitment_xlsx(records_or_session: Any, **kwargs: Any) -> bytes:
    """按腾讯文档兼容的招聘字段顺序导出 XLSX。"""

    kwargs.setdefault("columns", DEFAULT_RECRUITMENT_COLUMNS)
    return export_xlsx(records_or_session, **kwargs)


def save_export(data: bytes, path: str | Path) -> Path:
    """将导出字节保存到调用方指定路径，并返回绝对路径。"""

    target = Path(path).expanduser().resolve()
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(data)
    return target


def load_export_records(
    source: Any,
    *,
    statement: Any = None,
    model: Any = None,
    scalars: bool = True,
) -> list[Any]:
    """统一读取 list、SQLAlchemy Query 或 Session。"""

    if source is None:
        return []
    if statement is not None:
        if not hasattr(source, "execute"):
            raise TypeError("传入 statement 时 source 必须是 SQLAlchemy Session")
        result = source.execute(statement)
        return list(result.scalars().all() if scalars else result.mappings().all())
    if model is not None:
        if not hasattr(source, "query"):
            raise TypeError("传入 model 时 source 必须是 SQLAlchemy Session")
        return list(source.query(model).all())
    if hasattr(source, "all") and callable(source.all):
        return list(source.all())
    if isinstance(source, Iterable) and not isinstance(source, (str, bytes, Mapping)):
        return list(source)
    raise TypeError("请传入记录列表，或 Session + model/statement")


def record_to_dict(record: Any) -> dict[str, Any]:
    """将普通对象、dataclass、SQLAlchemy Row/模型转换为字典。"""

    if isinstance(record, Mapping):
        return dict(record)
    mapping = getattr(record, "_mapping", None)
    if mapping is not None:
        return dict(mapping)
    if is_dataclass(record) and not isinstance(record, type):
        return asdict(record)

    try:
        from sqlalchemy import inspect as sqlalchemy_inspect

        inspection = sqlalchemy_inspect(record)
        mapper = getattr(inspection, "mapper", None)
        if mapper is not None:
            return {
                attribute.key: getattr(record, attribute.key)
                for attribute in mapper.column_attrs
            }
    except (ImportError, TypeError, AttributeError):
        pass

    if hasattr(record, "__dict__"):
        return {
            key: value
            for key, value in vars(record).items()
            if not key.startswith("_")
        }
    raise TypeError(f"不支持导出的记录类型：{type(record).__name__}")


def _resolve_columns(
    records: Sequence[Any], columns: Any
) -> tuple[ExportColumn, ...]:
    if columns is None:
        keys: list[str] = []
        seen: set[str] = set()
        for record in records:
            for key in _flatten_mapping(record_to_dict(record)):
                if key not in seen:
                    seen.add(key)
                    keys.append(key)
        return tuple(ExportColumn(key, key) for key in keys)

    if isinstance(columns, Mapping):
        return tuple(
            ExportColumn(str(key), str(label)) for key, label in columns.items()
        )

    specs = []
    for item in columns:
        if isinstance(item, ExportColumn):
            specs.append(item)
        elif isinstance(item, str):
            specs.append(ExportColumn(item, item))
        elif isinstance(item, Mapping):
            key = str(item["key"])
            specs.append(
                ExportColumn(
                    key=key,
                    label=str(item.get("label", key)),
                    aliases=tuple(str(value) for value in item.get("aliases", ())),
                )
            )
        elif (
            isinstance(item, Sequence)
            and not isinstance(item, (str, bytes))
            and len(item) in (2, 3)
        ):
            aliases = tuple(str(value) for value in item[2]) if len(item) == 3 else ()
            specs.append(ExportColumn(str(item[0]), str(item[1]), aliases))
        else:
            raise TypeError("columns 项必须是字符串、映射、元组或 ExportColumn")
    return tuple(specs)


def _column_value(record: Any, column: ExportColumn) -> Any:
    paths = column.aliases or (column.key,)
    for path in paths:
        value = _get_path_value(record, path, _MISSING)
        if value is not _MISSING and value is not None:
            return value
    return None


def _get_path_value(record: Any, path: str, default: Any = None) -> Any:
    current = record
    for part in path.split("."):
        if isinstance(current, Mapping):
            if part not in current:
                return default
            current = current[part]
            continue
        mapping = getattr(current, "_mapping", None)
        if mapping is not None:
            if part not in mapping:
                return default
            current = mapping[part]
            continue
        if not hasattr(current, part):
            return default
        current = getattr(current, part)
    return current


def _flatten_mapping(
    value: Mapping[str, Any], prefix: str = ""
) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, item in value.items():
        path = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(item, Mapping):
            result.update(_flatten_mapping(item, path))
        else:
            result[path] = item
    return result


def _csv_value(value: Any) -> str:
    normalized = _normalized_value(value)
    return "" if normalized is None else str(normalized)


def _xlsx_value(value: Any) -> Any:
    normalized = _normalized_value(value)
    if normalized is None or isinstance(normalized, (str, int, float, bool)):
        return normalized
    return str(normalized)


def _normalized_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    if isinstance(value, Mapping):
        return json.dumps(
            _json_safe(value),
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        )
    if isinstance(value, Iterable) and not isinstance(value, (str, bytes)):
        return json.dumps(_json_safe(list(value)), ensure_ascii=False, separators=(",", ":"))
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _json_safe(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, Mapping):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, Iterable) and not isinstance(value, (str, bytes)):
        return [_json_safe(item) for item in value]
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _safe_sheet_name(value: str) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", str(value).strip()) or "招聘数据"
    return cleaned[:31]


def _display_width(value: Any) -> int:
    if value is None:
        return 0
    text = str(value)
    return min(
        max(
            (
                sum(2 if ord(character) > 127 else 1 for character in line)
                for line in text.splitlines() or [text]
            ),
            default=0,
        ),
        80,
    )


_MISSING = object()
