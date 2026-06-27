from __future__ import annotations

import io
import json
from collections.abc import Callable
from typing import Any

import pytest
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session, sessionmaker

from app.ai.client import LLMClient
from app.ai.prompts import SCORING_DIMENSIONS
from app.ai.service import AIService
from app.db import Base, create_db_engine
from app.models import Application, ApplicationStatus, Candidate, CandidateSource, User
from app.services.auth import hash_password, normalize_username, verify_password
from app.services.candidates import get_or_create_candidate, ingest_candidate
from app.services.exporter import export_recruitment_csv, export_recruitment_xlsx
from app.services.scheduling import (
    create_booking_token,
    generate_candidate_slots,
    intervals_conflict,
    validate_booking_token,
)
from app.services.workflow import (
    HumanActionRequired,
    InvalidStatusTransition,
    create_application,
    get_or_create_job_version,
    transition_application,
)


@pytest.fixture()
def db_session() -> Session:
    engine = create_db_engine("sqlite://")
    Base.metadata.create_all(engine)
    factory = sessionmaker(
        bind=engine,
        class_=Session,
        autoflush=False,
        expire_on_commit=False,
    )
    session = factory()
    try:
        yield session
    finally:
        session.close()
        Base.metadata.drop_all(engine)
        engine.dispose()


def _create_application(session: Session) -> Application:
    candidate, _ = get_or_create_candidate(
        session,
        {"name": "测试候选人", "phone": "13800138000"},
    )
    job, job_version, _ = get_or_create_job_version(
        session,
        {
            "platform": "test",
            "source_job_id": "job-1",
            "title": "后端工程师",
            "jd_raw": "负责 Python 和 FastAPI 服务开发",
        },
    )
    application, _ = create_application(
        session,
        {
            "candidate_id": candidate.id,
            "job_id": job.id,
            "job_version_id": job_version.id,
        },
    )
    return application


def test_candidate_deduplicates_by_normalized_name_and_phone(
    db_session: Session,
) -> None:
    first, first_created = get_or_create_candidate(
        db_session,
        {"name": " 张 三 ", "phone": "+86 138-0013-8000"},
    )
    second, second_created = get_or_create_candidate(
        db_session,
        {"name": "张三", "phone": "13800138000"},
    )

    assert first_created is True
    assert second_created is False
    assert second.id == first.id
    assert db_session.scalar(select(func.count()).select_from(Candidate)) == 1


def test_username_normalization_is_case_insensitive() -> None:
    assert normalize_username("  Admin  ") == "admin"
    assert normalize_username("TestUser") == normalize_username("testuser")


def test_password_hash_is_salted_and_verifiable() -> None:
    first_hash = hash_password("secure-password")
    second_hash = hash_password("secure-password")

    assert first_hash != second_hash
    assert "secure-password" not in first_hash
    assert verify_password("secure-password", first_hash) is True
    assert verify_password("wrong-password", first_hash) is False


def test_user_model_enforces_case_insensitive_identity(
    db_session: Session,
) -> None:
    user = User(
        username="Admin",
        normalized_username=normalize_username("Admin"),
        password_hash=hash_password("password123"),
    )
    db_session.add(user)
    db_session.flush()

    stored = db_session.scalar(
        select(User).where(
            User.normalized_username == normalize_username("admin")
        )
    )
    assert stored is not None
    assert stored.id == user.id


def test_missing_phone_ingest_is_idempotent_by_source(
    db_session: Session,
) -> None:
    candidate_data = {"name": "缺手机号候选人", "phone": None}
    source_data = {
        "platform": "mock-platform",
        "source_application_id": "application-without-phone-1",
        "raw_resume": "候选人暂未提供手机号",
    }

    first_candidate, first_source, first_created = ingest_candidate(
        db_session,
        candidate_data=candidate_data,
        source_data=source_data,
    )
    second_candidate, second_source, second_created = ingest_candidate(
        db_session,
        candidate_data=candidate_data,
        source_data=source_data,
    )

    assert first_created is True
    assert second_created is False
    assert second_candidate.id == first_candidate.id
    assert second_source.id == first_source.id
    assert db_session.scalar(select(func.count()).select_from(Candidate)) == 1
    assert db_session.scalar(select(func.count()).select_from(CandidateSource)) == 1


def test_illegal_application_status_transition_is_rejected(
    db_session: Session,
) -> None:
    application = _create_application(db_session)

    with pytest.raises(InvalidStatusTransition):
        transition_application(
            db_session,
            application.id,
            ApplicationStatus.REJECTED,
            actor="test",
            is_human_action=True,
        )

    assert application.status == ApplicationStatus.CAPTURED


def test_human_decision_status_requires_human_action(
    db_session: Session,
) -> None:
    application = _create_application(db_session)
    application.status = ApplicationStatus.EVALUATION_REVIEW
    db_session.flush()

    with pytest.raises(HumanActionRequired):
        transition_application(
            db_session,
            application.id,
            ApplicationStatus.INTERVIEW_PENDING,
            actor="evaluation-agent",
        )

    assert application.status == ApplicationStatus.EVALUATION_REVIEW
    transitioned = transition_application(
        db_session,
        application.id,
        ApplicationStatus.INTERVIEW_PENDING,
        actor="hr",
        is_human_action=True,
        reason="人工审批进入面试",
    )
    assert transitioned.status == ApplicationStatus.INTERVIEW_PENDING
    assert transitioned.decision_reason == "人工审批进入面试"


def test_mock_ai_returns_five_dimensions_and_correct_weighted_total() -> None:
    service = AIService(LLMClient(provider="mock"))
    result = service.evaluate_candidate(
        {
            "candidate": {
                "skills": ["Python", "FastAPI"],
                "project_experience": [
                    {"evidence": "使用 Python 和 FastAPI 开发 AI 招聘系统"}
                ],
                "internship_experience": [
                    {"evidence": "后端开发实习，负责接口和数据库"}
                ],
                "work_experience": [],
            },
            "chat": {
                "summary": "候选人有明确求职意向，可下周到岗",
                "intent": "愿意参加面试",
                "availability": "工作日下午均可",
            },
            "job": {
                "required_skills": ["Python", "FastAPI"],
                "experience_requirements": ["具有后端项目经验"],
                "domain": "AI",
            },
        }
    )

    expected_weights = {
        item["name"]: float(item["weight"]) for item in SCORING_DIMENSIONS
    }
    dimensions = result["dimensions"]

    assert len(dimensions) == 5
    assert {item["name"] for item in dimensions} == set(expected_weights)
    assert sum(item["weight"] for item in dimensions) == pytest.approx(1.0)
    for item in dimensions:
        assert item["weight"] == expected_weights[item["name"]]
        assert item["weighted_score"] == pytest.approx(
            round(item["score"] * item["weight"], 2)
        )
    assert result["total_score"] == pytest.approx(
        round(sum(item["score"] * item["weight"] for item in dimensions), 2)
    )


class _RecordingClient:
    def __init__(self) -> None:
        self.messages: list[dict[str, str]] = []

    def generate_json(
        self,
        *,
        messages: list[dict[str, str]],
        fallback: Callable[[], dict[str, Any]],
        validator: Callable[[dict[str, Any]], dict[str, Any]] | None = None,
    ) -> dict[str, Any]:
        self.messages = messages
        content = fallback()
        if validator is not None:
            content = validator(content)
        return {
            "content": content,
            "meta": {
                "requested_provider": "recording",
                "provider": "recording",
                "model": "recording",
                "fallback": False,
                "fallback_reason": None,
            },
        }


def test_public_ai_service_redacts_phone_before_model_call() -> None:
    client = _RecordingClient()
    service = AIService(client=client)  # type: ignore[arg-type]
    result = service.structure_candidate(
        resume_text="姓名：测试候选人\n手机号：13800138000\n技能：Python、FastAPI",
        chat_text="可以安排面试",
        jd_text="后端工程师，要求 Python 和 FastAPI",
    )

    serialized_messages = json.dumps(client.messages, ensure_ascii=False)
    assert "13800138000" not in serialized_messages
    assert "138****8000" in serialized_messages
    assert result["candidate"]["phone"] == "13800138000"


def test_generated_slots_do_not_conflict_with_busy_intervals() -> None:
    busy = [
        {
            "start_at": "2026-06-29T10:00:00+08:00",
            "end_at": "2026-06-29T11:00:00+08:00",
        }
    ]
    slots = generate_candidate_slots(
        "2026-06-29T09:00:00+08:00",
        "2026-06-29T12:00:00+08:00",
        busy,
        duration_minutes=60,
        step_minutes=30,
        buffer_minutes=0,
        limit=10,
    )

    assert [slot["start_at"].hour for slot in slots] == [9, 11]
    for slot in slots:
        assert not intervals_conflict(
            slot["start_at"],
            slot["end_at"],
            busy[0]["start_at"],
            busy[0]["end_at"],
        )


def test_generated_slots_accept_multiple_busy_interval_records() -> None:
    busy = [
        {
            "start_at": "2026-06-29T09:00:00+08:00",
            "end_at": "2026-06-29T10:00:00+08:00",
        },
        {
            "start_at": "2026-06-29T11:00:00+08:00",
            "end_at": "2026-06-29T12:00:00+08:00",
        },
    ]

    slots = generate_candidate_slots(
        "2026-06-29T09:00:00+08:00",
        "2026-06-29T13:00:00+08:00",
        busy,
        duration_minutes=60,
        step_minutes=60,
        buffer_minutes=0,
        limit=10,
    )

    assert [slot["start_at"].hour for slot in slots] == [10, 12]


def test_booking_token_signature_and_single_use_validation() -> None:
    secret = "unit-test-booking-secret"
    token = create_booking_token(
        {"interview_id": 42},
        secret,
        token_id="booking-token-42",
    )

    valid = validate_booking_token(token, secret)
    invalid = validate_booking_token(f"{token}tampered", secret)
    used = validate_booking_token(
        token,
        secret,
        used_token_ids={"booking-token-42"},
    )

    assert valid["ok"] is True
    assert valid["payload"]["interview_id"] == 42
    assert valid["payload"]["jti"] == "booking-token-42"
    assert invalid["ok"] is False
    assert invalid["code"] == "invalid_token"
    assert used["ok"] is False
    assert used["code"] == "token_used"


def test_csv_and_xlsx_exports_are_non_empty_and_readable() -> None:
    records = [
        {
            "candidate": {"name": "测试候选人"},
            "job": {"title": "后端工程师"},
            "application": {
                "status": "interview_pending",
                "updated_at": "2026-06-25T10:00:00+08:00",
            },
            "evaluation": {"total_score": 88.5},
            "owner": "HR",
            "interview": {"confirmed_slot": "2026-06-29T09:00:00+08:00"},
        }
    ]

    csv_data = export_recruitment_csv(records)
    xlsx_data = export_recruitment_xlsx(records)

    assert len(csv_data) > 0
    assert "测试候选人" in csv_data.decode("utf-8-sig")
    assert len(xlsx_data) > 0
    assert xlsx_data.startswith(b"PK")

    workbook = load_workbook(io.BytesIO(xlsx_data), read_only=True)
    worksheet = workbook.active
    assert worksheet.max_row == 2
    assert worksheet.cell(row=2, column=1).value == "测试候选人"
    workbook.close()
